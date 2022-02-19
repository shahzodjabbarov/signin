from pathlib import Path
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage
import tkinter as tk
import random
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

root = Tk()

root.geometry("700x500")
root.title("Moviesss")

filename=PhotoImage(file="files//signup.png")
button_image_1 = PhotoImage(file=("files//button.png"))
login_back = PhotoImage(file="files//loginpage.png")
login_button = PhotoImage(file="files//login.png")

background_label = Label()
entry_1 = Entry()
entry_2 = Entry()
button2 = Button()
button3 = Button()
button4 = Button()


def signup():
	
	global background_label
	global entry_1
	global entry_2
	global button2
	global button3
	global button4 
	entry_1.delete(0, tk.END)
	entry_2.delete(0, tk.END)
	background_label=Label(root,image=filename)
	background_label.place(x=0,y=0,relwidth=1,relheight=1)
	entry_1 = Entry(font = "Helvetica 13", fg="#013c3c", bd=0, bg="#ffffff", highlightthickness=0)
	entry_1.place(x=96.0, y=223.0, width=200.0, height=28.0)
	entry_2 = Entry(font = "Calibri 15 bold", fg="#013c3c", show="*",bd=0, bg="#ffffff", highlightthickness=0)
	entry_2.place(x=96.0, y=300.0, width=200.0, height=20.0) 
	button2 = Button(bg="#fbfafa",text="Already have an account?", font = "Calibri 11 underline", fg="#0f5054",borderwidth=0,highlightthickness=0,relief="flat", command=lambda:loginned())
	button2.place(x=94.0,y=410.0,width=200.0,height=31.0)
	button3 = Button(image=button_image_1,borderwidth=0,highlightthickness=0,relief="flat", command=lambda:checker())
	button3.place(x=96.0,y=355.0,width=200.0,height=31.0)





def loginned():
	
	global background_label
	global entry_1
	global entry_2
	global button2
	global button3
	global button4 
	entry_1.delete(0, tk.END)
	entry_2.delete(0, tk.END)
	background_label.configure(image=login_back)
	button3.configure(image=login_button, command=lambda:logger())
	button3.place(x=92.0,y=355.0,width=200.0,height=31.0)
	button2.destroy()
	button4 = Button(bg="#fbfafa",text="Don't have an account?", font = "Calibri 11 underline", fg="#0f5054",borderwidth=0,highlightthickness=0,relief="flat", command=lambda:signup())
	button4.place(x=95.0,y=410.0,width=200.0,height=31.0)

def checker():
	wb = load_workbook("files//users.xlsx") 
	ws = wb.active
	username = entry_1.get()
	password = entry_2.get()
	for i in range(2, ws.max_row+1):
		if (username == ws.cell(row=i, column=1).value):
			entry_1.delete(0, tk.END)
			entry_2.delete(0, tk.END)
			entry_1.insert(END, "Username already exists")
			return False
 

	else:
		ws.append([username, password])
		wb.save("files//users.xlsx")
		entry_1.delete(0, tk.END)
		entry_2.delete(0, tk.END)
		entry_1.insert(END, "Successfully signed up")



def logger():
	wb = load_workbook("files//users.xlsx") 
	ws = wb.active
	username = entry_1.get()
	password = entry_2.get()
	for i in range(2, ws.max_row+1):
		if (username == ws.cell(row=i, column=1).value):
			if (password == ws.cell(row=i, column=2).value):
				entry_1.delete(0, tk.END)
				entry_2.delete(0, tk.END)
				entry_1.insert(END, "Successfully logged in")
				break
			else:
				entry_1.delete(0, tk.END)
				entry_2.delete(0, tk.END)
				entry_1.insert(END, "Incorrect password")
				break
	else:
		entry_1.delete(0, tk.END)
		entry_2.delete(0, tk.END)
		entry_1.insert(END, "Username doesnt exist")
	



signup()

root.resizable(False, False)
root.mainloop()

